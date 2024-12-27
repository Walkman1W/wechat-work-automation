import pyautogui
import keyboard
import time
import pandas as pd
import pyperclip
from datetime import datetime
import logging
from typing import List, Dict
import os
from PIL import Image
import numpy as np
from mouse_recorder import MouseRecorder
import random

class MouseAutomation:
    def __init__(self):
        self.running = False
        self.paused = False
        self._setup_logging()
        self.mouse_recorder = MouseRecorder(self.logger)
        
    def _setup_logging(self):
        """设置日志"""
        # 创建logger
        self.logger = logging.getLogger('mouse_automation')
        self.logger.setLevel(logging.INFO)
        
        # 创建按日期命名的日志文件
        log_filename = f'automation_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
        file_handler = logging.FileHandler(log_filename, encoding='utf-8')
        
        # 设置日志格式
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        
        # 添加处理器
        self.logger.addHandler(file_handler)
        
        self.logger.info("程序启动")

    def record_coordinates(self, total_steps: int = None):
        """记录鼠标坐标的模块"""
        coordinates = self.mouse_recorder.record(total_steps)
        if coordinates:
            self.mouse_recorder.save_to_file()
        return True

    def _is_valid_phone(self, phone: str) -> bool:
        """验证手机号是否合法"""
        return len(str(phone)) == 11 and str(phone).isdigit()

    def _verify_template(self, x: int, y: int, template_path: str, threshold: float = 0.6, max_retries: int = 2, step_name: str = "", phone: str = "") -> tuple[bool, str]:
        """验证当前位置与模板是否匹配
        Args:
            x: 点击位置的x坐标
            y: 点击位置的y坐标
            template_path: 模板图片路径
            threshold: 匹配阈值，默认0.6
            max_retries: 最大重试次数，默认2次
            step_name: 步骤名称，用于保存失败截图
            phone: 手机号，用于保存失败截图
        Returns:
            (是否匹配, 失败时的截图路径)
        """
        retry_count = 0
        while retry_count < max_retries:
            try:
                # 检查模板文件是否存在
                if not os.path.exists(template_path):
                    self.logger.error(f"模板文件不存在: {template_path}")
                    print(f"模板文件不存在: {template_path}")
                    return False, ""
                    
                # 获取当前屏幕截图
                left = max(0, x - 40)
                top = max(0, y - 25)
                width = 80
                height = 50
                screenshot = pyautogui.screenshot(region=(left, top, width, height))
                
                # 加载模板图片
                template = Image.open(template_path)
                
                # 确保图片大小一致
                if screenshot.size != template.size:
                    self.logger.error(f"图片大小不匹配: 当前{screenshot.size} vs 模板{template.size}")
                    print(f"图片大小不匹配: 当前{screenshot.size} vs 模板{template.size}")
                    return False, ""
                
                # 转换为RGB模式并转为numpy数组
                screenshot_array = np.array(screenshot.convert('RGB'))
                template_array = np.array(template.convert('RGB'))
                
                # 分别计算RGB三个通道的相似度
                r_similarity = 1 - np.mean(np.abs(screenshot_array[:,:,0] - template_array[:,:,0]) / 255)
                g_similarity = 1 - np.mean(np.abs(screenshot_array[:,:,1] - template_array[:,:,1]) / 255)
                b_similarity = 1 - np.mean(np.abs(screenshot_array[:,:,2] - template_array[:,:,2]) / 255)
                
                # 计算总体相似度（三个通道的加权平均值，绿色通道权重更高）
                similarity = (0.3 * r_similarity + 0.4 * g_similarity + 0.3 * b_similarity)
                
                # 计算局部区域相似度（将图像分成16个区域）
                h, w = screenshot_array.shape[:2]
                h_step = h // 4
                w_step = w // 4
                local_similarities = []
                
                for i in range(4):
                    for j in range(4):
                        h_start = i * h_step
                        h_end = (i + 1) * h_step if i < 3 else h
                        w_start = j * w_step
                        w_end = (j + 1) * w_step if j < 3 else w
                        
                        region_screenshot = screenshot_array[h_start:h_end, w_start:w_end]
                        region_template = template_array[h_start:h_end, w_start:w_end]
                        
                        region_similarity = 1 - np.mean(np.abs(region_screenshot - region_template) / 255)
                        local_similarities.append(region_similarity)
                
                # 最终相似度是全局相似度和局部相似度的加权平均，更重视局部相似度
                final_similarity = 0.4 * similarity + 0.6 * min(local_similarities)
                print(f"最终相似度: {final_similarity:.4f}")
                
                # 打印匹配结果
                if final_similarity >= threshold:
                    print("模板匹配成功")
                    return True, ""
                
                print("模板匹配失败")
                if retry_count < max_retries - 1:
                    print(f"等待3秒后重试...")
                    time.sleep(3)
                    retry_count += 1
                    continue
                
                # 最后一次失败时保存截图
                debug_path = f'debug_screenshots/{step_name}_{phone}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
                os.makedirs('debug_screenshots', exist_ok=True)
                screenshot.save(debug_path)
                print(f"失败截图已保存至: {debug_path}")
                
                # 记录详细的匹配信息到日志文件
                self.logger.debug(f"""模板匹配详细信息:
                    模板文件: {os.path.basename(template_path)}
                    当前位置: ({x}, {y})
                    全局相似度: {similarity:.4f}
                    最低局部相似度: {min(local_similarities):.4f}
                    最终相似度: {final_similarity:.4f}
                    匹配阈值: {threshold}
                    当前重试次数: {retry_count + 1}/{max_retries}
                """)
                return False, debug_path
                
            except Exception as e:
                self.logger.error(f"模板验证失败: {e}")
                print(f"模板验证失败: {e}")
                print(f"错误类型: {type(e)}")
                if retry_count < max_retries - 1:
                    print(f"等待3秒后重试...")
                    time.sleep(3)
                    retry_count += 1
                    continue
                return False, ""
        
        return False, ""

    def _print_summary(self, df):
        """打印自动化处理结果统计
        Args:
            df: 处理后的DataFrame
        """
        print("\n自动化处理结果统计")
        print("=" * 50)
        
        # 统计各种状态的数量
        total_records = len(df)
        processed = len(df[df['状态'] == '已处理'])
        failed = len(df[df['状态'] == '添加失败'])
        invalid = len(df[df['状态'] == '无效手机号'])
        error = len(df[df['状态'].str.startswith('错误:', na=False)]) if '状态' in df else 0
        skipped = len(df[pd.isna(df['状态'])])
        
        # 计算成功率
        success_rate = (processed / total_records * 100) if total_records > 0 else 0
        
        # 打印统计结果
        print(f"总记录数: {total_records}")
        print(f"处理成功: {processed} ({success_rate:.1f}%)")
        print(f"添加失败: {failed}")
        print(f"无效号码: {invalid}")
        print(f"发生错误: {error}")
        print(f"未处理数: {skipped}")
        print("=" * 50)
        
        # 记录到日志
        self.logger.info(f"""自动化处理结果统计:
            总记录数: {total_records}
            处理成功: {processed} ({success_rate:.1f}%)
            添加失败: {failed}
            无效号码: {invalid}
            发生错误: {error}
            未处理数: {skipped}
        """)

    def _get_random_delay(self, min_delay=3, max_delay=6):
        """生成随机延迟时间
        Args:
            min_delay: 最小延迟秒数
            max_delay: 最大延迟秒数
        Returns:
            随机延迟秒数
        """
        return random.uniform(min_delay, max_delay)

    def automate_process(self):
        """自动化处理模块"""
        print("\n开始自动化处理...")
        self.logger.info("开始自动化处理")
        
        # 添加连续失败计数器
        consecutive_failures = 0
        
        # 加载Excel数据
        try:
            print("正在加载Excel文件...")
            df = pd.read_excel('phone.xlsx')
            print(f"成功加载Excel文件，共 {len(df)} 条记录")
            self.logger.info(f"成功加载Excel文件，共 {len(df)} 条记录")
        except Exception as e:
            error_msg = f"加载Excel文件失败: {e}"
            print(error_msg)
            self.logger.error(error_msg)
            return True

        print("正在加载坐标文件...")
        coordinates = self.mouse_recorder.load_from_file()
        
        if not coordinates:
            self.logger.warning("未找到坐标文件或坐标文件为空")
            print("未找到坐标文件或坐标文件为空，请先记录坐标")
            return True
            
        # 检查坐标点数量是否足够
        if len(coordinates) < 5:  # 需要5个坐标点
            error_msg = f"坐标点数量不足，需要5个坐标点，当前只有{len(coordinates)}个"
            print(error_msg)
            self.logger.error(error_msg)
            return True
        
        print(f"成功加载坐标文件，共 {len(coordinates)} 个坐标点")
        self.logger.info(f"成功加载坐标文件，共 {len(coordinates)} 个坐标点")
        
        # 注册快捷键
        keyboard.add_hotkey('ctrl+f1', self._toggle_pause)
        keyboard.add_hotkey('ctrl+f2', self._stop)
        
        self.running = True
        print("\n开始自动化处理，按Ctrl+F1暂停/继续，按Ctrl+F2结束")
        print("=" * 50)

        try:
            for index, row in df.iterrows():
                if not self.running:
                    print("\n检测到停止信号，结束处理")
                    self.logger.info("检测到停止信号，结束处理")
                    break

                # 如果状态不为空，跳过
                if pd.notna(row['状态']):
                    print(f"跳过已处理的记录: {row['手机号']}")
                    self.logger.info(f"跳过已处理的记录: {row['手机号']}")
                    continue

                # 验证手机号
                if not self._is_valid_phone(row['手机号']):
                    print(f"无效的手机号: {row['手机号']}")
                    self.logger.warning(f"无效的手机号: {row['手机号']}")
                    df.at[index, '状态'] = '无效手机号'
                    continue

                while self.paused:
                    time.sleep(0.1)

                try:
                    print(f"\n正在处理第 {index + 1} 条记录，手机号: {row['手机号']}")
                    self.logger.info(f"开始处理第 {index + 1} 条记录，手机号: {row['手机号']}")
                    
                    # 步骤延时设置（现在作为基础延时）
                    base_delays = {
                        'step1': 2,  # 点击添加按钮后等待
                        'step2': 3,  # 点击输入框后等待
                        'step3': 3,  # 点击添加按钮后等待
                        'step4': 3,  # 点击发送邀请后等待
                        'step5': 3   # 点击确认后等待
                    }
                    
                    # Step 1: 点击添加
                    print("步骤1: 点击添加按钮")
                    self.logger.debug("执行步骤1: 点击添加按钮")
                    # 添加随机延迟
                    random_delay = self._get_random_delay()
                    print(f"等待 {random_delay:.1f} 秒...")
                    time.sleep(random_delay)
                    
                    x, y = coordinates[0]['step1']['x'], coordinates[0]['step1']['y']
                    success, debug_path = self._verify_template(x, y, coordinates[0]['step1']['template'], step_name="step1", phone=str(row['手机号']))
                    if not success:
                        error_msg = f"步骤1验证失败：界面不匹配 {debug_path}"
                        print(error_msg)
                        self.logger.error(error_msg)
                        df.at[index, '状态'] = '添加失败'
                        consecutive_failures += 1
                        
                        # 检查连续失败次数
                        if consecutive_failures >= 2:
                            print("\n警告：检测到连续2次匹配失败！")
                            print("请检查以下可能的问题：")
                            print("1. 微信窗口是否被遮挡或最小化")
                            print("2. 界面是否发生变化")
                            print("3. 坐标点是否需要重新记录")
                            print("\n正在返回主菜单...")
                            self.logger.warning("检测到连续2次匹配失败，自动返回主菜单")
                            return True
                            
                        # 立即保存Excel
                        df.to_excel('phone.xlsx', index=False)
                        # 设置列宽
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook('phone.xlsx')
                            ws = wb.active
                            ws.column_dimensions['A'].width = 9
                            ws.column_dimensions['B'].width = 9
                            ws.column_dimensions['C'].width = 15
                            ws.column_dimensions['D'].width = 15
                            wb.save('phone.xlsx')
                        except Exception as e:
                            self.logger.warning(f"设置Excel列宽失败: {e}")
                        continue
                    
                    # 重置连续失败计数器
                    consecutive_failures = 0
                    pyautogui.click(x=x, y=y)
                    time.sleep(base_delays['step1'])

                    # Step 2: 点击输入框并输入手机号
                    print("步骤2: 点击输入框并输入手机号")
                    self.logger.debug("执行步骤2: 点击输入框并输入手机号")
                    # 添加随机延迟
                    random_delay = self._get_random_delay()
                    print(f"等待 {random_delay:.1f} 秒...")
                    time.sleep(random_delay)
                    
                    x, y = coordinates[1]['step2']['x'], coordinates[1]['step2']['y']
                    success, debug_path = self._verify_template(x, y, coordinates[1]['step2']['template'], step_name="step2", phone=str(row['手机号']))
                    if not success:
                        error_msg = f"步骤2验证失败：界面不匹配 {debug_path}"
                        print(error_msg)
                        self.logger.error(error_msg)
                        df.at[index, '状态'] = '添加失败'
                        consecutive_failures += 1
                        
                        # 检查连续失败次数
                        if consecutive_failures >= 2:
                            print("\n警告：检测到连续2次匹配失败！")
                            print("请检查以下可能的问题：")
                            print("1. 微信窗口是否被遮挡或最小化")
                            print("2. 界面是否发生变化")
                            print("3. 坐标点是否需要重新记录")
                            print("\n正在返回主菜单...")
                            self.logger.warning("检测到连续2次匹配失败，自动返回主菜单")
                            return True
                            
                        # 立即保存Excel
                        df.to_excel('phone.xlsx', index=False)
                        # 设置列宽
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook('phone.xlsx')
                            ws = wb.active
                            ws.column_dimensions['A'].width = 9
                            ws.column_dimensions['B'].width = 9
                            ws.column_dimensions['C'].width = 15
                            ws.column_dimensions['D'].width = 15
                            wb.save('phone.xlsx')
                        except Exception as e:
                            self.logger.warning(f"设置Excel列宽失败: {e}")
                        continue
                    
                    # 重置连续失败计数器
                    consecutive_failures = 0
                    pyautogui.click(x=x, y=y)
                    time.sleep(base_delays['step2'])
                    pyperclip.copy(str(row['手机号']))
                    # 模拟手动输入的随机延迟
                    time.sleep(random.uniform(0.5, 1.5))
                    pyautogui.hotkey('ctrl', 'v')
                    time.sleep(random.uniform(0.3, 0.8))
                    pyautogui.press('enter')
                    time.sleep(1.0)

                    # Step 3: 点击添加按钮
                    print("步骤3: 点击添加按钮")
                    self.logger.debug("执行步骤3: 点击添加按钮")
                    # 添加随机延迟
                    random_delay = self._get_random_delay()
                    print(f"等待 {random_delay:.1f} 秒...")
                    time.sleep(random_delay)
                    
                    x, y = coordinates[2]['step3']['x'], coordinates[2]['step3']['y']
                    success, debug_path = self._verify_template(x, y, coordinates[2]['step3']['template'], step_name="step3", phone=str(row['手机号']))
                    if not success:
                        error_msg = f"步骤3验证失败：界面不匹配 {debug_path}"
                        print(error_msg)
                        self.logger.error(error_msg)
                        df.at[index, '状态'] = '添加失败'
                        consecutive_failures += 1
                        
                        # 检查连续失败次数
                        if consecutive_failures >= 2:
                            print("\n警告：检测到连续2次匹配失败！")
                            print("请检查以下可能的问题：")
                            print("1. 微信窗口是否被遮挡或最小化")
                            print("2. 界面是否发生变化")
                            print("3. 坐标点是否需要重新记录")
                            print("\n正在返回主菜单...")
                            self.logger.warning("检测到连续2次匹配失败，自动返回主菜单")
                            return True
                            
                        # 立即保存Excel
                        df.to_excel('phone.xlsx', index=False)
                        # 设置列宽
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook('phone.xlsx')
                            ws = wb.active
                            ws.column_dimensions['A'].width = 9
                            ws.column_dimensions['B'].width = 9
                            ws.column_dimensions['C'].width = 15
                            ws.column_dimensions['D'].width = 15
                            wb.save('phone.xlsx')
                        except Exception as e:
                            self.logger.warning(f"设置Excel列宽失败: {e}")
                        continue
                    
                    # 重置连续失败计数器
                    consecutive_failures = 0
                    pyautogui.click(x=x, y=y)
                    time.sleep(base_delays['step3'])

                    # Step 4: 点击发送邀请
                    print("步骤4: 点击发送邀请")
                    self.logger.debug("执行步骤4: 点击发送邀请")
                    # 添加随机延迟
                    random_delay = self._get_random_delay()
                    print(f"等待 {random_delay:.1f} 秒...")
                    time.sleep(random_delay)
                    
                    x, y = coordinates[3]['step4']['x'], coordinates[3]['step4']['y']
                    success, debug_path = self._verify_template(x, y, coordinates[3]['step4']['template'], step_name="step4", phone=str(row['手机号']))
                    if not success:
                        error_msg = f"步骤4验证失败：界面不匹配 {debug_path}"
                        print(error_msg)
                        self.logger.error(error_msg)
                        df.at[index, '状态'] = '添加失败'
                        consecutive_failures += 1
                        
                        # 检查连续失败次数
                        if consecutive_failures >= 2:
                            print("\n警告：检测到连续2次匹配失败！")
                            print("请检查以下可能的问题：")
                            print("1. 微信窗口是否被遮挡或最小化")
                            print("2. 界面是否发生变化")
                            print("3. 坐标点是否需要重新记录")
                            print("\n正在返回主菜单...")
                            self.logger.warning("检测到连续2次匹配失败，自动返回主菜单")
                            return True
                            
                        # 立即保存Excel
                        df.to_excel('phone.xlsx', index=False)
                        # 设置列宽
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook('phone.xlsx')
                            ws = wb.active
                            ws.column_dimensions['A'].width = 9
                            ws.column_dimensions['B'].width = 9
                            ws.column_dimensions['C'].width = 15
                            ws.column_dimensions['D'].width = 15
                            wb.save('phone.xlsx')
                        except Exception as e:
                            self.logger.warning(f"设置Excel列宽失败: {e}")
                        continue
                    
                    # 重置连续失败计数器
                    consecutive_failures = 0
                    pyautogui.click(x=x, y=y)
                    time.sleep(base_delays['step4'])

                    # Step 5: 点击确认按钮
                    print("步骤5: 点击确认按钮")
                    self.logger.debug("执行步骤5: 点击确认按钮")
                    # 添加随机延迟
                    random_delay = self._get_random_delay()
                    print(f"等待 {random_delay:.1f} 秒...")
                    time.sleep(random_delay)
                    
                    x, y = coordinates[4]['step5']['x'], coordinates[4]['step5']['y']
                    success, debug_path = self._verify_template(x, y, coordinates[4]['step5']['template'], step_name="step5", phone=str(row['手机号']))
                    if not success:
                        error_msg = f"步骤5验证失败：界面不匹配 {debug_path}"
                        print(error_msg)
                        self.logger.error(error_msg)
                        df.at[index, '状态'] = '添加失败'
                        consecutive_failures += 1
                        
                        # 检查连续失败次数
                        if consecutive_failures >= 2:
                            print("\n警告：检测到连续2次匹配失败！")
                            print("请检查以下可能的问题：")
                            print("1. 微信窗口是否被遮挡或最小化")
                            print("2. 界面是否发生变化")
                            print("3. 坐标点是否需要重新记录")
                            print("\n正在返回主菜单...")
                            self.logger.warning("检测到连续2次匹配失败，自动返回主菜单")
                            return True
                            
                        # 立即保存Excel
                        df.to_excel('phone.xlsx', index=False)
                        # 设置列宽
                        try:
                            from openpyxl import load_workbook
                            wb = load_workbook('phone.xlsx')
                            ws = wb.active
                            ws.column_dimensions['A'].width = 9
                            ws.column_dimensions['B'].width = 9
                            ws.column_dimensions['C'].width = 15
                            ws.column_dimensions['D'].width = 15
                            wb.save('phone.xlsx')
                        except Exception as e:
                            self.logger.warning(f"设置Excel列宽失败: {e}")
                        continue
                    
                    # 重置连续失败计数器
                    consecutive_failures = 0
                    pyautogui.click(x=x, y=y)
                    print(f"手机号 {row['手机号']} 处理完成")
                    self.logger.info(f"手机号 {row['手机号']} 处理完成")
                    df.at[index, '状态'] = '已处理'
                    time.sleep(base_delays['step5'])

                except Exception as e:
                    error_msg = f"处理手机号 {row['手机号']} 时出错: {e}"
                    print(error_msg)
                    self.logger.error(error_msg)
                    df.at[index, '状态'] = f'错误: {str(e)}'

                # 保存进度
                df.to_excel('phone.xlsx', index=False)
                
                # 设置列宽
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook('phone.xlsx')
                    ws = wb.active
                    # 设置列宽：第1,2列为9，第3,4列为15
                    ws.column_dimensions['A'].width = 9
                    ws.column_dimensions['B'].width = 9
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    wb.save('phone.xlsx')
                except Exception as e:
                    self.logger.warning(f"设置Excel列宽失败: {e}")
                
                print("进度已保存到Excel文件")
                self.logger.info("进度已保存到Excel文件")
                print("-" * 50)

        except Exception as e:
            error_msg = f"自动化处理出错: {e}"
            print(error_msg)
            self.logger.error(error_msg)
        finally:
            # 清理快捷键
            keyboard.unhook_all()
            print("\n自动化处理完成")
            self.logger.info("自动化处理完成")
            print("=" * 50)
            
            # 打印处理结果统计
            self._print_summary(df)
            
            return True

    def _toggle_pause(self):
        """暂停/继续自动化处理"""
        self.paused = not self.paused
        status = "已暂停" if self.paused else "继续运行"
        print(status)
        self.logger.info(status)

    def _stop(self):
        """停止自动化处理"""
        self.running = False
        self.logger.info("程序已停止")
        print("程序已停止")

def main():
    automation = MouseAutomation()
    while True:
        print("\n请选择模式:")
        print("1: 记录坐标")
        print("2: 自动化处理")
        print("3: 退出程序")
        
        mode = input("请输入选择 (1/2/3): ")
        automation.logger.info(f"用户选择模式: {mode}")
        
        if mode == "1":
            steps = input("请输入需要记录的坐标数量（直接回车则不限制数量）: ")
            total_steps = int(steps) if steps.strip() else None
            automation.record_coordinates(total_steps)
        elif mode == "2":
            automation.automate_process()
        elif mode == "3":
            automation.logger.info("程序退出")
            print("程序已退出")
            break
        else:
            print("无效的选择，请重新输入")
            automation.logger.warning(f"无效的模式选择: {mode}")

if __name__ == "__main__":
    main() 